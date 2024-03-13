import openpyxl
import redis


def read_excel_and_write_to_redis(file_path, redis_db):
    # Open the Excel file
    wb = openpyxl.load_workbook(file_path)
    # metadataSheet = wb.active

    # 处理任务元数据
    metadataSheet = wb['metadata']
    handleMetaData(metadataSheet, redis_db)

    # 处理全局任务
    globalTaskSheet = wb['globalTask']
    handleGlobalTask(globalTaskSheet, redis_db)

    # 处理任务池
    taskPoolSheet = wb['taskPool']
    handleTaskPool(taskPoolSheet, redis_db)


def handleMetaData(sheet, redis_db):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 接收数据
        app_id, task_group_id, period_id, object_id, task_id, progress, claim_prog, task_group_finish, replace_count, replace_list = row
        # 拼接 key
        key = f"evt:mt:{int(app_id)}:{int(task_group_id)}:{int(period_id)}:{{{object_id}}}"

        # 设置任务进度
        prField = f"pr_{int(task_id)}"
        redis_db.hset(key, prField, int(progress))

        # 设置领奖状态
        rsField = f"rs_{int(task_id)}"
        redis_db.hset(key, rsField, int(claim_prog))

        # 设置任务组完成标记
        if task_group_finish:
            finishField = "group_finish"
            redis_db.hset(key, finishField, "true")


def handleGlobalTask(sheet, redis_db):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 接收数据
        app_id, object_id, task_id, progress = row
        # 拼接 key
        key = f"evt:global:{int(app_id)}:{int(task_id)}:{{{object_id}}}"
        # 写库
        redis_db.set(key, int(progress))


def handleTaskPool(sheet, redis_db):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 接收数据
        app_id, task_group_id, period_id, object_id, task_pool = row
        # 拼接 key
        poolKey = f"evt:tk:{int(app_id)}:{int(task_group_id)}:{int(period_id)}:{{{object_id}}}"
        # 切分数据
        task_list = task_pool.split(',')
        # 写库
        redis_db.sadd(poolKey, *task_list)


if __name__ == "__main__":
    # Connect to Redis
    r = redis.Redis(host='10.0.1.71', port=6379, db=7)
    #
    # Read Excel and write to Redis
    read_excel_and_write_to_redis("transfer.xlsx", r)

    print("process end")

